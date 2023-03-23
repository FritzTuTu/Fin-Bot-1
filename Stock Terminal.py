import certifi
import json
import openpyxl
from urllib.request import urlopen
from prettytable import PrettyTable

def get_jsonparsed_data(url):
    """
    Receive the content of ``url``, parse it as JSON and return the object.

    Parameters
    ----------
    url : str

    Returns
    -------
    dict
    """
    response = urlopen(url, cafile=certifi.where())
    data = response.read().decode("utf-8")
    return json.loads(data)

# Define a function to get the stock code from the user
def get_stock_code():
    stock_code = input("Enter the stock code: ")
    return stock_code

# Define a function to ask the user if they want to save the data to an Excel file
def ask_to_save():
    while True:
        save = input("Do you want to save the data to an Excel file? (Y/N): ")
        if save.lower() == "y":
            return True
        elif save.lower() == "n":
            return False
        else:
            print("Invalid input. Please enter Y or N.")

# Call the get_stock_code function to get the user input
stock_code = get_stock_code()

url = f"https://financialmodelingprep.com/api/v3/ratios/{stock_code}?apikey=6773f916da46be1f54613e01994d476d"
data = get_jsonparsed_data(url)

# Define a function to display data in a table format
def display_data(data):
    table = PrettyTable()
    # Add columns to the table
    table.field_names = list(data[0].keys())
    # Add rows to the table
    for row in data:
        table.add_row(list(row.values()))
    print(table)

# Call the display_data function to print the data in a table
display_data(data)

# Ask the user if they want to save the data to an Excel file
save_data = ask_to_save()

if save_data:
    # Write the data to an Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    # Add header row
    header = list(data[0].keys())
    for col in range(1, len(header)+1):
        ws.cell(row=1, column=col).value = header[col-1]
    # Add data rows
    for row in range(2, len(data)+2):
        for col in range(1, len(header)+1):
            ws.cell(row=row, column=col).value = list(data[row-2].values())[col-1]
    # Save the workbook
    filename = f"{stock_code}.xlsx"
    wb.save(filename)
    print(f"Data saved to {filename}.")
else:
    print("Data not saved.")


